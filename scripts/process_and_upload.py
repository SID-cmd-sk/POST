"""
SKS Scantech — GitHub Actions Processor
========================================
Runs automatically when a new ZIP lands in submissions/.
  1. Extracts data.json from each ZIP
  2. Writes a row to database.xlsx (on Google Drive)
  3. Uploads the ZIP itself into Drive /processed folder
  4. Moves the ZIP in the repo from submissions/ → processed/

Secrets required (set in GitHub repo → Settings → Secrets):
  GDRIVE_CREDENTIALS  : content of service-account JSON key file
  GDRIVE_FOLDER_ID    : Google Drive folder ID where database.xlsx lives
"""

import os, json, zipfile, shutil, base64, tempfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import io

# ── Constants ──────────────────────────────────────────────────────────────
SCOPES       = ['https://www.googleapis.com/auth/drive']
SUBMISSIONS  = Path('submissions')
PROCESSED    = Path('processed')
DB_FILENAME  = 'database.xlsx'

HEADERS    = ["Date", "Machine Name", "Machine Type", "Axis",
              "Controller", "Email", "Contact"]
COL_WIDTHS = [14, 28, 18, 14, 20, 28, 18]

THIN        = Side(border_style="thin", color="DDDDDD")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Google Drive helpers ────────────────────────────────────────────────────
def get_drive_service():
    creds_json = os.environ['GDRIVE_CREDENTIALS']
    creds_info = json.loads(creds_json)
    creds = service_account.Credentials.from_service_account_info(
        creds_info, scopes=SCOPES)
    return build('drive', 'v3', credentials=creds)


def find_file_in_drive(service, folder_id, filename):
    """Return file ID if filename exists in the Drive folder, else None."""
    q = (f"'{folder_id}' in parents and name='{filename}' "
         f"and trashed=false")
    results = service.files().list(q=q, fields='files(id,name)').execute()
    files = results.get('files', [])
    return files[0]['id'] if files else None


def download_db_from_drive(service, folder_id, local_path):
    """Download database.xlsx from Drive to local_path. Return True if found."""
    file_id = find_file_in_drive(service, folder_id, DB_FILENAME)
    if not file_id:
        return False
    request = service.files().get_media(fileId=file_id)
    with open(local_path, 'wb') as f:
        downloader = MediaIoBaseDownload(f, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    return True


def upload_file_to_drive(service, folder_id, local_path, mime, replace=False):
    """Upload or update a file in Drive. Returns file ID."""
    filename = os.path.basename(local_path)
    media    = MediaFileUpload(local_path, mimetype=mime, resumable=True)
    existing_id = find_file_in_drive(service, folder_id, filename)

    if existing_id and replace:
        updated = service.files().update(
            fileId=existing_id, media_body=media).execute()
        return updated['id']
    else:
        meta = {'name': filename, 'parents': [folder_id]}
        created = service.files().create(
            body=meta, media_body=media, fields='id').execute()
        return created['id']


# ── Excel helpers (same style as your original script) ─────────────────────
def style_header_row(ws):
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        c.fill      = PatternFill("solid", fgColor="CC0000")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = CELL_BORDER
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS[i-1]
    ws.row_dimensions[1].height = 22


def style_data_row(ws, row_idx):
    fill = PatternFill("solid", fgColor="FFF5F5" if row_idx % 2 == 0 else "FFFFFF")
    for i in range(1, len(HEADERS)+1):
        c = ws.cell(row=row_idx, column=i)
        c.fill      = fill
        c.font      = Font(name="Calibri", size=10)
        c.alignment = Alignment(vertical="center")
        c.border    = CELL_BORDER


def get_or_create_sheet(wb, customer_name):
    safe = "".join(ch for ch in customer_name
                   if ch not in r'\/:*?[]')[:31].strip() or "Unknown"
    if safe in wb.sheetnames:
        return wb[safe], False
    ws = wb.create_sheet(title=safe)
    style_header_row(ws)
    ws.freeze_panes = "A2"
    return ws, True


def existing_machine_names(ws):
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[1] is not None:
            names.add((row[1] or "").strip().lower())
    return names


def append_machine_row(ws, today_str, machine, email, contact):
    r = ws.max_row + 1
    ws.cell(r, 1, today_str)
    ws.cell(r, 2, machine.get("machine_name", ""))
    ws.cell(r, 3, machine.get("machine_type", ""))
    ws.cell(r, 4, machine.get("axis", ""))
    ws.cell(r, 5, machine.get("controller", ""))
    ws.cell(r, 6, email)
    ws.cell(r, 7, contact)
    style_data_row(ws, r)


def write_log_sheet(wb, records):
    sn = "_Log"
    if sn in wb.sheetnames:
        ws = wb[sn]
    else:
        ws = wb.create_sheet(title=sn)
        hdrs = ["File", "Customer Name", "Status", "Timestamp", "Notes"]
        for i, h in enumerate(hdrs, 1):
            c = ws.cell(1, i, h)
            c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            c.fill      = PatternFill("solid", fgColor="1A1A1A")
            c.alignment = Alignment(horizontal="center")
        for i, w in enumerate([30, 25, 12, 20, 60], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    color_map = {"Processed": "00AA00", "Duplicate": "FF8800", "Error": "CC0000"}
    for rec in records:
        r = ws.max_row + 1
        ws.cell(r, 1, rec["file"])
        ws.cell(r, 2, rec["customer_name"])
        sc = ws.cell(r, 3, rec["status"])
        sc.font = Font(bold=True,
                       color=color_map.get(rec["status"], "000000"),
                       name="Calibri", size=10)
        ws.cell(r, 4, rec["timestamp"])
        ws.cell(r, 5, rec["notes"])


# ── Main processing loop ────────────────────────────────────────────────────
def main():
    folder_id = os.environ['GDRIVE_FOLDER_ID']
    service   = get_drive_service()
    today_str = date.today().strftime("%Y-%m-%d")

    PROCESSED.mkdir(exist_ok=True)

    # Find all unprocessed ZIPs
    zip_files = sorted(SUBMISSIONS.glob("*.zip"))
    if not zip_files:
        print("No new ZIPs found in submissions/. Exiting.")
        return

    # Download current database.xlsx from Drive (or start fresh)
    with tempfile.TemporaryDirectory() as tmpdir:
        db_local = os.path.join(tmpdir, DB_FILENAME)
        found    = download_db_from_drive(service, folder_id, db_local)
        if found:
            wb = load_workbook(db_local)
            print(f"Downloaded existing {DB_FILENAME} from Drive.")
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            print(f"{DB_FILENAME} not found in Drive — creating new.")

        records = []

        for zip_path in zip_files:
            filename  = zip_path.name
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            record    = dict(file=filename, customer_name="",
                             status="Error", timestamp=timestamp, notes="")
            try:
                if not zipfile.is_zipfile(zip_path):
                    raise ValueError("Not a valid ZIP file.")

                with zipfile.ZipFile(zip_path) as zf:
                    json_names = [n for n in zf.namelist()
                                  if os.path.basename(n).lower() == "data.json"]
                    if not json_names:
                        raise FileNotFoundError("data.json not found inside ZIP.")
                    raw  = zf.open(json_names[0]).read().decode("utf-8-sig")
                    data = json.loads(raw)

                customer_name = (data.get("customer_name") or "").strip()
                email         = (data.get("email")         or "").strip()
                contact       = (data.get("contact")       or "").strip()
                machines      = data.get("machines", [])

                if not customer_name:
                    raise ValueError("customer_name is empty in data.json.")

                record["customer_name"] = customer_name
                print(f"  Processing: {filename}  →  {customer_name}")

                ws, is_new = get_or_create_sheet(wb, customer_name)
                existing   = existing_machine_names(ws)
                added, skipped, notes_arr = 0, 0, []

                for m in machines:
                    m_name = (m.get("machine_name") or "").strip()
                    if not m_name:
                        skipped += 1
                        continue
                    if m_name.lower() in existing:
                        notes_arr.append(f"Duplicate: {m_name}")
                        skipped += 1
                    else:
                        append_machine_row(ws, today_str, m, email, contact)
                        existing.add(m_name.lower())
                        added += 1

                record["status"] = "Processed"
                record["notes"]  = f"Added {added}, skipped {skipped}. " + "; ".join(notes_arr)
                print(f"  ✓ Added {added} machine(s), skipped {skipped}.")

                # Upload ZIP to Drive /processed sub-folder
                upload_file_to_drive(
                    service, folder_id,
                    str(zip_path),
                    'application/zip'
                )
                print(f"  ✓ ZIP uploaded to Drive.")

                # Move ZIP in repo: submissions/ → processed/
                dest = PROCESSED / filename
                if dest.exists():
                    base, ext = os.path.splitext(filename)
                    dest = PROCESSED / f"{base}_{timestamp.replace(':','-')}{ext}"
                shutil.move(str(zip_path), str(dest))
                print(f"  ✓ ZIP moved to processed/.")

            except Exception as e:
                record["status"] = "Error"
                record["notes"]  = str(e)
                print(f"  ✗ ERROR: {filename} — {e}")

            records.append(record)

        # Write log sheet and save workbook
        write_log_sheet(wb, records)
        wb.save(db_local)
        print(f"\nSaving updated {DB_FILENAME}…")

        # Upload updated database.xlsx back to Drive (replace existing)
        upload_file_to_drive(
            service, folder_id,
            db_local,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            replace=True
        )
        print(f"✓ {DB_FILENAME} uploaded to Google Drive successfully.")

    n_ok  = sum(1 for r in records if r["status"] == "Processed")
    n_err = sum(1 for r in records if r["status"] == "Error")
    print(f"\nDone — Processed: {n_ok}  |  Errors: {n_err}")


if __name__ == "__main__":
    main()
